# -*- coding: utf-8 -*-
import scrapy
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook

class Test2Spider(scrapy.Spider):
    name = 'test2'
    allowed_domains = ['www.greenresidential.com/owners/owner-faqs/']
    start_urls = ['http://www.greenresidential.com/owners/owner-faqs//']

    def parse(self, response):
        wb = Workbook()
        ws = wb.active   
        qn = response.xpath('//*[@class="acc_head"]/text()').extract()
        ans = response.xpath('//*[@class="acc_content"]/p').extract()

        count = 1 
        for i in qn:
            print("Did quesiton run?")
            ws["A"+str(count)] = BeautifulSoup(i, "lxml").text
            count+=1
        count = 1
        for j in ans:
            ws["B"+str(count)] = BeautifulSoup(j, "lxml").text
            count+=1
  
        wb.save("test2.xlsx")

        '''yield {'Question':qn,
                'Answer': ans}'''
