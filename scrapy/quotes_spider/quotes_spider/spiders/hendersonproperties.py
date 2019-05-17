# -*- coding: utf-8 -*-
import scrapy
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook

class HendersonpropertiesSpider(scrapy.Spider):
    name = 'hendersonproperties'
    allowed_domains = ['www.hendersonproperties.com/property-management-services/faqs/']
    start_urls = ['http://www.hendersonproperties.com/property-management-services/faqs//']


    def parse(self, response):
        wb = Workbook()
        ws = wb.active   
 
        qn = response.xpath('//*[@class="title"]').extract()
        ans = response.xpath('//*[@class="answer"]').extract()

        count = 1 
        for i in qn:
            print("Did quesiton run?")
            ws["A"+str(count)] = BeautifulSoup(i, "lxml").text
            count+=1
        count = 1
        for j in ans:
            ws["B"+str(count)] = BeautifulSoup(j, "lxml").text
            count+=1
        wb.save("hendersonproperties.xlsx")


