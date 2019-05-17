# -*- coding: utf-8 -*-
import scrapy
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook

class QuotesSpider(scrapy.Spider):
    name = 'quotes'
    allowed_domains = ['www.sutd.edu.sg/Admissions/Undergraduate/FAQs']
    start_urls = ['http://www.sutd.edu.sg/Admissions/Undergraduate/FAQs/']

    def parse(self, response):
        wb = Workbook()
        ws = wb.active   
 
        qn = response.xpath('//*[@class="accordion-title"]/h2').extract()
        ans = response.xpath('//*[@class="accordion-content"]').extract()
        #excelWrite('A',qn)
        #excelWrite('B',qn)
        count = 1 
        for i in qn:
            print("Did quesiton run?")
            ws["A"+str(count)] = BeautifulSoup(i, "lxml").text
            count+=1
        count = 1
        for j in ans:
            ws["B"+str(count)] = BeautifulSoup(j, "lxml").text
            count+=1
        wb.save("trial3.xlsx")


