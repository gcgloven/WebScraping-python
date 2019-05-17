# -*- coding: utf-8 -*-
import scrapy
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook

class FaqSpider(scrapy.Spider):
    name = 'faq'
    allowed_domains = ['pm1sf.com/faq']
    start_urls = ['http://pm1sf.com/faq/']

    def parse(self, response):
        wb = Workbook()
        ws = wb.active   
        faq = response.xpath('//*[@id="block-017fcf8e00b3d1bf8307"]/div/p').extract()
        #print (faq)
        
        count = 1
        i=0
        while (i < len(faq)):
            print("what is happening ?????????????????????????????????????/")
            ws["A"+str(count)] = BeautifulSoup(faq[i], "lxml").text
            i+=1
            ws["B"+str(count)] = BeautifulSoup(faq[i], "lxml").text
            i+=1
            count+=1
        
        wb.save("faq1.xlsx")
	
